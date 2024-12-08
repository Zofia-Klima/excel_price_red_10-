import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename, price_column, corrected_price_column, chart_placement_cell):
    wb = xl.load_workbook(filename)
    sheet1 = wb["Sheet1"]


    for row in range(2, sheet1.max_row + 1):
        cell = sheet1.cell(row, price_column)
        corrected_price = round(cell.value * 0.9, 2)
        corrected_price_cell = sheet1.cell(row, corrected_price_column)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet1,
                       min_row=2,
                       max_row=sheet1.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet1.add_chart(chart, chart_placement_cell)

    wb.save(filename)